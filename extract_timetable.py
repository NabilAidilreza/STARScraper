import os
from bs4 import BeautifulSoup
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill, Font

### HTML READER FUNCTIONS ###

def read_html_file(file_name):
    # Get the current directory of the script
    script_directory = os.path.dirname(os.path.abspath(__file__))
    # Construct the file path by joining the script directory and the file name
    file_path = os.path.join(script_directory, file_name)
    # Open the file in read mode and specify the encoding as utf-8
    with open(file_path, 'r', encoding='windows-1252') as file:
        # Read the content of the file
        html_content = file.read()
        # Return the HTML content
        return html_content

def process_html_to_data(FILE_NAME):
    # Set the name of the HTML file to be processed
    file_name = FILE_NAME
    # Read the contents of the HTML file
    html_content = read_html_file(file_name)
    # Create an empty list to store the table data
    table = []
    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')
    # Find the main content section of the HTML
    mainContent = soup.find_all('table')[3]
    # Find all the rows in the main content section
    content_Rows = mainContent.find_all('tr')
    # Iterate over each row in the content_Rows list
    for row in content_Rows:
        # Find all the data cells in the current row
        data = row.find_all('td')
        # Create an empty list to store the processed row data
        temp_row = []
        # Iterate over each data cell in the current row
        for d in data:
            # Process the text by removing newlines
            process_text = d.text.replace('\n','')
            # Replace any '\xa0' with an empty string
            if process_text == '\xa0':
                process_text = ''
            # Append the processed text to the temp_row list
            temp_row.append(process_text)
        # Append the temp_row list (excluding the first element) to the table list
        table.append(temp_row[1:])

    # Fix the last part of the table
    table[-1][0],table[-1][1] = table[-1][1],table[-1][0]
    table[-1][0] = "Total AU Registered"
    table[-1][-1] = ""

    # Return the table data
    print("Reading HTML...")
    # Output raw data in list form #
    return table

### --------------------- ###

### HELPER SORTING FUNCTIONS ###

def get_day_number(day):
    # Define a dictionary to map full weekday names to their numeric representation
    days_mapping = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
    # Extract the abbreviated weekday name from the full name (assuming the first 3 characters represent the abbreviation)
    abbreviated_day = day[:3]
    return days_mapping.get(abbreviated_day,-1)

def get_week_from_remark(remark):
    # Remove the 'Teaching Wk' prefix from the remark string
    week_str = remark.replace('Teaching Wk', '').strip()
    # Check if the remaining string is a digit
    if week_str.isdigit():
        # If it is, return a list with that digit as the only element
        return [int(week_str)]
    else:
        # If it's not a digit, return 0
        return 0

### ------------------------ ###

def write_to_csv():
    table = process_html_to_data()
    with open('Courses.csv', 'w',newline="") as f:
        # create the csv writer
        writer = csv.writer(f)
        # write a row to the csv file
        writer.writerows(table)

def process_data(FILE_NAME):
    # Prepare Data #
    file_name = FILE_NAME
    table = []
    table = process_html_to_data(file_name)
    print("Converting into sorted array...")
    # Fill up blanks #
    for i in range(len(table)):
        for m in range(len(table[i])):
            if table[i][m] == '':
                table[i][m] = table[i-1][m]

    # Standarize Wk Symbols #
    for i in range(len(table)):
        for m in range(len(table[i])):
            if "," in  table[i][m]:
                table[i][m] = table[i][m].replace(",","-")

    # Clean Last Row #
    table[-1][0],table[-1][1] = table[-1][1],table[-1][0]
    table[-1][0] = "Total AU Registered"
    table[-1][-1] = ""

    course_info = table

    # Clean table #
    course_info = [course for course in course_info if all(course)]

    # Sort by Day #
    sorted_array = sorted(course_info, key=lambda x: (get_day_number(x[11]), x[12], get_week_from_remark(x[14])))

    # Returns a sorted array based on raw list provided #
    return sorted_array

def further_process_data(table):
    course_table = table
    extra_table = []
    to_duplicate = []
    extract_table = course_table[1:]
    print("Preparing data for timetable format...")
    # Create duplicates for sorting by week purposes #
    for i in range(len(extract_table)):
        temp = extract_table[i][14].strip("Teaching Wk").split("-")
        if len(temp) != 1:
            to_duplicate.append([i,int(temp[0]),int(temp[1])])

    # Adjust duplicates #
    for d in to_duplicate:
        to_add = [extract_table[d[0]]] * (d[2]-d[1]+1)
        temp = []
        for m in range(d[1],d[2]+1):
            temp_course = extract_table[d[0]]
            temp.append(temp_course[:14] + ["Teaching Wk"+str(m)] + temp_course[15:])
        extra_table.append(temp)

    main_table = []
    
    # Add single week courses, ignore original duplicates #
    counter = 0
    while len(to_duplicate) > 0:
        if counter == to_duplicate[0][0]:
            del to_duplicate[0]
        elif counter >= len(extract_table):
            break
        else:
            main_table.append(extract_table[counter])
        counter += 1

    # Combine all list #
    for e in extra_table:
        for m in e:
            main_table.append(m)

    course_info = main_table

    # Clean table #
    course_info = [course for course in course_info if all(course)]
    
    # Sort by Week -> Day -> Time #
    sorted_array = sorted(course_info, key=lambda x: (get_week_from_remark(x[14]), get_day_number(x[11]), x[12]))
    sorted_array.insert(0,course_table[0])

    # Break up into different sets (by week) #
    curr = ''
    prev = ''
    final_array = []
    same_week_list = []
    for array in sorted_array:
        curr = array[14]
        if curr != prev:
            prev = curr
            final_array.append(same_week_list)
            same_week_list = []
        same_week_list.append(array)
    if same_week_list != []:
        final_array.append(same_week_list)
        same_week_list = []

    # Clear empty list at the start #
    del final_array[0]
    print("Finalizing final array...")
    return final_array

def write_timetable_to_csv(data,raw_data):
    #headers = data[0][0]
    headers = ["","Title","ModuleNo","Time","Venue","Group","ClassType","IndexNo","AU","CourseType","S/U","GERType","Status","Choice","Remark","Exam"]
    data = data[1:]
    print("Writing data into Excel...")
    workbook = openpyxl.Workbook()

    ### Create Overview Sheet ###
    workbook.create_sheet(title="Overview")
    sheet = workbook["Overview"]
    for row_idx, row in enumerate(raw_data, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Auto-adjust column width
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2  # Adding some buffer space and adjusting
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Auto-adjust row height
    for row_cells in sheet.rows:
        max_height = 0
        for cell in row_cells:
            if cell.value is not None:
                lines = str(cell.value).count('\n') + 1
                max_height = max(max_height, lines)
        adjusted_height = max_height * 15  # You can adjust the row height as needed
        sheet.row_dimensions[row_cells[0].row].height = adjusted_height


    ### Create Wk 1-13 Sheets ###
    n=13
    week_list = [f'Wk{i}' for i in range(1, n+1)]
    for week_name in week_list:
        workbook.create_sheet(title=week_name)

    # Write data into weeks sheets respectively #
    for week_number, week_name in enumerate(week_list, start=1):
        sheet = workbook[week_name]
        # Write headers to the first row
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        for row_idx, row in enumerate(data[week_number-1], start=2):
            for col_idx, cell_value in enumerate(row, start=1):
                #sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                #order = [12,2,1,13,14,11,10,7,3,4,5,6,8,9,15,16]
                if col_idx == 12:
                    sheet.cell(row=row_idx, column=1, value=cell_value)
                elif col_idx == 2:
                    sheet.cell(row=row_idx, column=2, value=cell_value)
                elif col_idx == 1:
                    sheet.cell(row=row_idx, column=3, value=cell_value)
                elif col_idx == 13:
                    sheet.cell(row=row_idx, column=4, value=cell_value)
                elif col_idx == 14:
                    sheet.cell(row=row_idx, column=5, value=cell_value)
                elif col_idx == 11:
                    sheet.cell(row=row_idx, column=6, value=cell_value)
                elif col_idx == 10:
                    sheet.cell(row=row_idx, column=7, value=cell_value)
                elif col_idx == 7:
                    sheet.cell(row=row_idx, column=8, value=cell_value)

                elif col_idx == 3:
                    sheet.cell(row=row_idx, column=9, value=cell_value)
                elif col_idx == 4:
                    sheet.cell(row=row_idx, column=10, value=cell_value)
                elif col_idx == 5:
                    sheet.cell(row=row_idx, column=11, value=cell_value)
                elif col_idx == 6:
                    sheet.cell(row=row_idx, column=12, value=cell_value)
                elif col_idx == 8:
                    sheet.cell(row=row_idx, column=13, value=cell_value)
                elif col_idx == 9:
                    sheet.cell(row=row_idx, column=14, value=cell_value)
                elif col_idx == 15:
                    sheet.cell(row=row_idx, column=15, value=cell_value)
                elif col_idx == 16:
                    sheet.cell(row=row_idx, column=16, value=cell_value)
                
                # Course No 1
                # Title 2
                # AU 3
                # CourseType 4
                # S/U Grade option 5
                # GERType 6
                # IndexNumber 7
                # Status 8
                # Choice 9
                # ClassType 10
                # Group 11
                # Day 12
                # Time 13
                # Venue 14
                # Remark 15
                # Exam 16


        # Optionally, you can set a title for the sheet
        sheet['A1'] = f"Day"

        # Auto-adjust column width
        for column_cells in sheet.columns:
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2  # Adding some buffer space and adjusting
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Auto-adjust row height
        for row_cells in sheet.rows:
            max_height = 0
            for cell in row_cells:
                if cell.value is not None:
                    lines = str(cell.value).count('\n') + 1
                    max_height = max(max_height, lines)
            adjusted_height = max_height * 15  # You can adjust the row height as needed
            sheet.row_dimensions[row_cells[0].row].height = adjusted_height

    # You can remove the default first sheet 'Sheet' if you don't need it
    workbook.remove(workbook['Sheet'])
    workbook.save("weekly_data.xlsx")
                
def color_cells(workbook):
    # Color Coding #
    color_Mon = '538DD5' # light blue
    color_Tue = '76933C' # light 
    color_Wed = 'FABF8F' #
    color_Thurs = 'B1A0C7' #
    color_Fri = '366092' #
    n = 13
    week_list = [f'Wk{i}' for i in range(1, n+1)]

    worksheet = workbook["Overview"]
    for row in worksheet: 
        if row[11].value == "Day":
            for cell in row:
                cell.font = Font(bold=True)
        if row[11].value == "Mon":
            clr_background = PatternFill(start_color='5eb5e9', end_color='5eb5e9', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Tue":
            clr_background = PatternFill(start_color='4cc249', end_color='4cc249', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Wed":
            clr_background = PatternFill(start_color='dba658', end_color='dba658', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Thu":
            clr_background = PatternFill(start_color='eeeb59', end_color='eeeb59', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Fri":
            clr_background = PatternFill(start_color='0f67b7', end_color='0f67b7', fill_type="solid")
            for cell in row:
                cell.fill = clr_background

    for week in week_list:
        worksheet = workbook[week]
        for row in worksheet: 
            if row[0].value == "Day":
                for cell in row:
                    cell.font = Font(bold=True)
            if row[0].value == "Mon":
                clr_background = PatternFill(start_color='5eb5e9', end_color='5eb5e9', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Tue":
                clr_background = PatternFill(start_color='4cc249', end_color='4cc249', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Wed":
                clr_background = PatternFill(start_color='dba658', end_color='dba658', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Thu":
                clr_background = PatternFill(start_color='eeeb59', end_color='eeeb59', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Fri":
                clr_background = PatternFill(start_color='0f67b7', end_color='0f67b7', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background

    # Save the changes back to the Excel file
    workbook.save("weekly_data.xlsx")


def see_table(final_data):
    for d in final_data:
        for e in d:
            print(e)

def create_excel_timetable(FILE_NAME):
    sorted_data = process_data(FILE_NAME)
    temp = sorted_data
    final_data = further_process_data(sorted_data)
    # Write to Excel #
    write_timetable_to_csv(final_data,temp)
    # Organize in Excel #
    workbook = openpyxl.load_workbook("weekly_data.xlsx")
    color_cells(workbook)

def create_timetable_list(FILE_NAME):
    sorted_data = process_data(FILE_NAME)
    temp = sorted_data
    final_data = further_process_data(sorted_data)
    return final_data




