import csv
import os
from datetime import datetime, timedelta

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill

# FILE THAT MANAGES ALL BASIC FUNCTIONS OF PROJECT #

#? Named column indices for a parsed course row (16 columns, after slicing the radio cell) #
COURSE_CODE = 0
COURSE_TITLE = 1
AU = 2
COURSE_TYPE = 3
SU_OPTION = 4
GER_TYPE = 5
INDEX_NUMBER = 6
STATUS = 7
CHOICE = 8
CLASS_TYPE = 9
GROUP = 10
DAY = 11
TIME = 12
VENUE = 13
REMARK = 14
EXAM = 15
NUM_COLUMNS = 16

NUM_TEACHING_WEEKS = 13
# NTU schedule: 13 teaching weeks + 1 recess week. Recess always falls after teaching week 7,
# so teaching weeks 1-7 map to calendar weeks 1-7, and weeks 8-13 map to calendar weeks 9-14.
RECESS_AFTER_WEEK = 7

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DAY_INDEX = {day: idx for idx, day in enumerate(DAY_ORDER)}

### Date Helpers ###

def parse_start_date(start_date):
    """Parse a DD/MM/YYYY string into a datetime (start of the day)."""
    day, month, year = start_date.split("/")
    return datetime(int(year), int(month), int(day), 0, 0, 0)


def teaching_week_start(teaching_week, start_date):
    """Return the Monday datetime of the given teaching week (1-13).

    Accounts for the recess week (after teaching week 7): weeks 1-7 are offset by
    (week - 1), weeks 8-13 are offset by (week) to skip the recess week.
    """
    startday = parse_start_date(start_date)
    if teaching_week <= RECESS_AFTER_WEEK:
        week_offset = teaching_week - 1
    else:
        week_offset = teaching_week
    return startday + timedelta(weeks=week_offset)


### HTML Reader Functions ###
# Opens and read html content from html file #
def read_html_file(file_name):
    script_directory = os.path.dirname(os.path.abspath(__file__))
    main_directory = os.path.dirname(script_directory)
    file_path = os.path.join(main_directory, file_name)
    for encoding in ('utf-8', 'windows-1252'):
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                return file.read()
        except UnicodeDecodeError:
            continue
    with open(file_path, 'rb') as file:
        return file.read().decode('utf-8', errors='replace')


# Convert html content to python list #
def process_html_to_data(FILE_NAME):
    html_content = read_html_file(FILE_NAME)
    table = []
    soup = BeautifulSoup(html_content, 'html.parser')
    mainContent = _find_timetable_table(soup)
    content_Rows = mainContent.find_all('tr')
    for row in content_Rows:
        data = row.find_all('td')
        temp_row = []
        for d in data:
            process_text = d.text.replace('\n', '').replace('\xa0', '')
            temp_row.append(process_text)
        if not temp_row:
            continue
        # Skip the "Total AU Registered" footer row entirely (label may sit in either
        # of the first two cells depending on the STARS layout)
        if any(cell.strip() == "Total AU Registered" for cell in temp_row[:2]):
            continue
        table.append(temp_row[1:])
    # Drop any fully-empty rows
    table = [row for row in table if row]
    return table


def _find_timetable_table(soup):
    """Locate the registered-courses table robustly instead of hardcoding table index [3].

    The courses table is the one whose last row begins with "Total AU Registered".
    """
    for row in soup.find_all('tr'):
        cells = [cell.text.replace('\n', '').strip() for cell in row.find_all('td')]
        if cells and cells[0] == "Total AU Registered":
            return row.find_parent('table')
    # Fallback: the historical layout put the course list at index 3
    tables = soup.find_all('table')
    if len(tables) > 3:
        return tables[3]
    raise ValueError("Could not locate the course timetable table in the HTML.")


### -------------------------###

### Helper Sorting Functions ###

# Sort by day helper func #
def get_day_number(day):
    # Returns 0-6 for Mon-Sun, -1 for unknown
    return DAY_INDEX.get(day, -1)


# Sort by week helper func #
def get_week_from_remark(remark):
    # Remove the 'Teaching Wk' prefix from the remark string
    week_str = remark.replace('Teaching Wk', '').strip()
    if week_str.isdigit():
        return int(week_str)
    return 0


# Expand a week remark like "Teaching Wk1-13" / "Teaching Wk2,6" into a list of week numbers #
def expand_weeks(remark):
    spec = remark.replace('Teaching Wk', '').strip()
    weeks = []
    for part in spec.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            try:
                start, end = part.split('-')
                weeks.extend(range(int(start), int(end) + 1))
            except ValueError:
                continue
        else:
            try:
                weeks.append(int(part))
            except ValueError:
                continue
    return weeks


### -------------------------###

### Data Filtering Functions ###

# Settle blanks, standarization, general cleaning, sorting #
def process_data(FILE_NAME):
    # Prepare Data #
    table = process_html_to_data(FILE_NAME)
    # Fill up blanks (skip the header row to avoid wrapping to the last row) #
    for i in range(1, len(table)):
        for m in range(min(len(table[i]), len(table[i - 1]))):
            if table[i][m] == '':
                table[i][m] = table[i - 1][m]
    # Clean table #
    course_info = [course for course in table if all(course)]
    # Sort by Day #
    sorted_array = sorted(course_info, key=lambda x: (get_day_number(x[DAY]), x[TIME]))
    # Returns a sorted array based on raw list provided #
    return sorted_array


# Settle duplicates, filling missing weeks, final cleaning and sorting #
def further_process_data(table):
    course_table = table
    extract_table = course_table[1:]
    # Create a per-week duplicate for every teaching week the course runs #
    main_table = []
    for row in extract_table:
        weeks = expand_weeks(row[REMARK])
        if not weeks:
            weeks = list(range(1, NUM_TEACHING_WEEKS + 1))
        for week in weeks:
            main_table.append(row[:REMARK] + ["Teaching Wk" + str(week)] + row[REMARK + 1:])
    # Clean table #
    course_info = [course for course in main_table if all(course)]
    # Sort by Week -> Day -> Time #
    sorted_array = sorted(
        course_info,
        key=lambda x: (get_week_from_remark(x[REMARK]), get_day_number(x[DAY]), x[TIME]),
    )
    sorted_array.insert(0, course_table[0])
    # Break up into different sets (by week), keeping the header row as the first group #
    final_array = []
    curr = None
    same_week_list = []
    for array in sorted_array:
        if array[REMARK] != curr:
            if same_week_list:
                final_array.append(same_week_list)
            same_week_list = []
            curr = array[REMARK]
        same_week_list.append(array)
    if same_week_list:
        final_array.append(same_week_list)
    return final_array


# Creates a python list, fully sorted and cleaned #
def create_timetable_list(FILE_NAME):
    """Create the sorted/cleaned timetable from a STAR Planner HTML file.

    Auto-detects the format: the usual planner (rows of registered courses with
    a 'Total AU Registered' footer) or the weekly grid view (TIME\\DAY grid +
    an '@Exam Schedule' course list). Both produce the same per-week structure.
    """
    if _is_weekly_format(FILE_NAME):
        from .ntu_weekly_format_extract import create_weekly_timetable_list
        try:
            return create_weekly_timetable_list(FILE_NAME)
        except (ValueError, IndexError, KeyError):
            pass  # misclassified; fall back to the standard parser below
    sorted_data = process_data(FILE_NAME)
    final_data = further_process_data(sorted_data)
    return final_data


def _is_weekly_format(FILE_NAME):
    """Sniff whether the HTML is the weekly grid export rather than the usual planner.

    Weekly files carry an '@Exam Schedule' course-list header and never contain
    the standard 'Total AU Registered' footer; the usual planner is the reverse.
    (Both formats contain a TIME\\DAY grid, so that alone cannot discriminate.)
    """
    content = read_html_file(FILE_NAME)
    return "@exam schedule" in content.lower() and "total au registered" not in content.lower()

### -------------------------###

### Timeline Functions ###

def generate_timeline(start_date):
    # Returns one list of Mon-Sat dates per teaching week (13 entries) #
    timeline = []
    for week in range(1, NUM_TEACHING_WEEKS + 1):
        week_start = teaching_week_start(week, start_date)
        week_dates = [(week_start + timedelta(days=j)).strftime("%d/%m/%Y") for j in range(6)]
        timeline.append(week_dates)
    return timeline

# CSV / Excel Writing Functions ###

def write_timetable_to_csv(data, raw_data):
    #headers = data[0][0]
    headers = ["","Title","ModuleNo","Time","Venue","Group","ClassType","IndexNo","AU","CourseType","S/U","GERType","Status","Choice","Remark","Exam"]
    data = data[1:]
    print("Writing data into Excel...")
    workbook = openpyxl.Workbook()
    ### Create Overview Sheet ###
    workbook.create_sheet(title="Overview")
    sheet = workbook["Overview"]
    for row_idx, row in enumerate(raw_data, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    # Auto-adjust column width
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2  # Adding some buffer space and adjusting
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    # Auto-adjust row height
    for row_cells in sheet.rows:
        max_height = 0
        for cell in row_cells:
            if cell.value is not None:
                lines = str(cell.value).count('\n') + 1
                max_height = max(max_height, lines)
        adjusted_height = max_height * 15  # You can adjust the row height as needed
        sheet.row_dimensions[row_cells[0].row].height = adjusted_height
    ### Create Wk 1-13 Sheets ###
    n=13
    week_list = [f'Wk{i}' for i in range(1, n+1)]
    for week_name in week_list:
        workbook.create_sheet(title=week_name)
    # Write data into weeks sheets respectively #
    for week_number, week_name in enumerate(week_list, start=1):
        sheet = workbook[week_name]
        # Write headers to the first row
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        for row_idx, row in enumerate(data[week_number-1], start=2):
            for col_idx, cell_value in enumerate(row, start=1):
                if col_idx == 12:
                    sheet.cell(row=row_idx, column=1, value=cell_value)
                elif col_idx == 2:
                    sheet.cell(row=row_idx, column=2, value=cell_value)
                elif col_idx == 1:
                    sheet.cell(row=row_idx, column=3, value=cell_value)
                elif col_idx == 13:
                    sheet.cell(row=row_idx, column=4, value=cell_value)
                elif col_idx == 14:
                    sheet.cell(row=row_idx, column=5, value=cell_value)
                elif col_idx == 11:
                    sheet.cell(row=row_idx, column=6, value=cell_value)
                elif col_idx == 10:
                    sheet.cell(row=row_idx, column=7, value=cell_value)
                elif col_idx == 7:
                    sheet.cell(row=row_idx, column=8, value=cell_value)
                elif col_idx == 3:
                    sheet.cell(row=row_idx, column=9, value=cell_value)
                elif col_idx == 4:
                    sheet.cell(row=row_idx, column=10, value=cell_value)
                elif col_idx == 5:
                    sheet.cell(row=row_idx, column=11, value=cell_value)
                elif col_idx == 6:
                    sheet.cell(row=row_idx, column=12, value=cell_value)
                elif col_idx == 8:
                    sheet.cell(row=row_idx, column=13, value=cell_value)
                elif col_idx == 9:
                    sheet.cell(row=row_idx, column=14, value=cell_value)
                elif col_idx == 15:
                    sheet.cell(row=row_idx, column=15, value=cell_value)
                elif col_idx == 16:
                    sheet.cell(row=row_idx, column=16, value=cell_value)
        # Optionally, you can set a title for the sheet
        sheet['A1'] = "Day"
        # Auto-adjust column width
        for column_cells in sheet.columns:
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2  # Adding some buffer space and adjusting
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        # Auto-adjust row height
        for row_cells in sheet.rows:
            max_height = 0
            for cell in row_cells:
                if cell.value is not None:
                    lines = str(cell.value).count('\n') + 1
                    max_height = max(max_height, lines)
            adjusted_height = max_height * 15
            sheet.row_dimensions[row_cells[0].row].height = adjusted_height
    workbook.remove(workbook['Sheet'])
    workbook.save("weekly_data.xlsx")

def write_to_csv(FILE_NAME):
    table = process_html_to_data(FILE_NAME)
    with open('Courses.csv', 'w', newline="") as f:
        # create the csv writer
        writer = csv.writer(f)
        # write a row to the csv file
        writer.writerows(table)

def color_cells(workbook):
    # Color Coding #
    n = 13
    week_list = [f'Wk{i}' for i in range(1, n+1)]
    worksheet = workbook["Overview"]
    for row in worksheet: 
        if row[11].value == "Day":
            for cell in row:
                cell.font = Font(bold=True)
        if row[11].value == "Mon":
            clr_background = PatternFill(start_color='5eb5e9', end_color='5eb5e9', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Tue":
            clr_background = PatternFill(start_color='4cc249', end_color='4cc249', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Wed":
            clr_background = PatternFill(start_color='dba658', end_color='dba658', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Thu":
            clr_background = PatternFill(start_color='eeeb59', end_color='eeeb59', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
        if row[11].value == "Fri":
            clr_background = PatternFill(start_color='0f67b7', end_color='0f67b7', fill_type="solid")
            for cell in row:
                cell.fill = clr_background
    for week in week_list:
        worksheet = workbook[week]
        for row in worksheet: 
            if row[0].value == "Day":
                for cell in row:
                    cell.font = Font(bold=True)
            if row[0].value == "Mon":
                clr_background = PatternFill(start_color='5eb5e9', end_color='5eb5e9', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Tue":
                clr_background = PatternFill(start_color='4cc249', end_color='4cc249', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Wed":
                clr_background = PatternFill(start_color='dba658', end_color='dba658', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Thu":
                clr_background = PatternFill(start_color='eeeb59', end_color='eeeb59', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
            if row[0].value == "Fri":
                clr_background = PatternFill(start_color='0f67b7', end_color='0f67b7', fill_type="solid")
                for cell in row:
                    cell.fill = clr_background
    # Save the changes back to the Excel file
    workbook.save("weekly_data.xlsx")

def create_excel_timetable(FILE_NAME):
    sorted_data = process_data(FILE_NAME)
    temp = sorted_data
    final_data = further_process_data(sorted_data)
    # Write to Excel #
    write_timetable_to_csv(final_data,temp)
    # Organize in Excel #
    workbook = openpyxl.load_workbook("weekly_data.xlsx")
    color_cells(workbook)
